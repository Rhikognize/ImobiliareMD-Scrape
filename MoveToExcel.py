from openpyxl import Workbook
from ScrapingParsing import get_info, sleep

count = 2

wb = Workbook()

ws = wb.active

ws["A1"] = "Title"
ws["B1"] = "Location"
ws["C1"] = "Rooms"
ws["D1"] = "Shower Rooms"
ws["E1"] = "Area"
ws["F1"] = "Type"
ws["G1"] = "Status"
ws["H1"] = "Price"
ws["I1"] = "Living"
ws["J1"] = "Repair"
ws["K1"] = "Construction"
ws["L1"] = "FLoor"
ws["M1"] = "Land"
ws["N1"] = "URL"

for i in get_info():
    ws[f"A{count}"] = i[0]
    ws[f"B{count}"] = i[1]
    ws[f"C{count}"] = i[2]
    ws[f"D{count}"] = i[3]
    ws[f"E{count}"] = i[4]
    ws[f"F{count}"] = i[5]
    ws[f"G{count}"] = i[6]
    ws[f"H{count}"] = i[7]
    ws[f"I{count}"] = i[8]
    ws[f"J{count}"] = i[9]
    ws[f"K{count}"] = i[10]
    ws[f"L{count}"] = i[11]
    ws[f"M{count}"] = i[12]
    ws[f"N{count}"] = i[13]
    count += 1
    try:
        wb.save("Real estate data.xlsx")
    except:
        pass

print("All data was extracted to Excel")
